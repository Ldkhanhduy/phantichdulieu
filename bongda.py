import pandas as pd
import numpy as n
import matplotlib.pyplot as plt
import openpyxl as o

#Bảng 1
def bieu_do_cot1(ex):
    danh_gia = ["Kém", "Trung bình", "Khá", "Tốt"]
    dk = [0, 2, 3, 4, 5]
    ex["Xếp loại"] = pd.cut(ex["Consumer_Rating"], bins=dk, labels=danh_gia)
    print("Bảng tổng lượt đánh giá qua các năm theo mức độ xếp loại")
    ga = pd.crosstab(index=ex["Year"], columns=ex["Xếp loại"])
    print(ga)
    ga_1 = ga
    ga_1.plot.bar()
    plt.xlabel("Năm")
    plt.ylabel("Lượt đánh giá")
    plt.title("Biểu đồ lượt đánh giá qua các năm 2019-2022")
    plt.grid(True)
    plt.show()
    df = pd.read_excel("D:/Duy/honda_sample_2.xlsx", sheet_name="Sheet1")
    merged_df = pd.concat([df, ga], ignore_index=False)
    book = o.load_workbook("D:/Duy/honda_sample_2.xlsx")
    writer = pd.ExcelWriter("D:/Duy/honda_sample_2.xlsx", engine='openpyxl')
    writer.book = book
    merged_df.to_excel(writer, sheet_name="Sheet1", index=False)
    writer.save()

#Bảng 2
def bieu_do_duong(ex):
    ga_1 = pd.crosstab(index=ex["Year"], columns=ex["Fuel_Type"])
    ga_1[["Gasoline", "Hybrid"]].plot(marker="*")
    print("Bảng tổng số xe bán được theo loại động cơ qua các năm")
    print(ga_1)
    plt.title("Biểu đồ thể hiện số xe bán được theo hai loại máy nhiên liệu qua các năm")
    plt.ylabel("Số lượng")
    plt.grid(True)
    plt.show()
    df = pd.read_excel("D:/Duy/honda_sample_2.xlsx", sheet_name="Sheet2")
    merged_df = pd.concat([df, ga_1], ignore_index=False)
    book = o.load_workbook("D:/Duy/honda_sample_2.xlsx")
    writer = pd.ExcelWriter("D:/Duy/honda_sample_2.xlsx", engine='openpyxl')
    writer.book = book
    merged_df.to_excel(writer, sheet_name="Sheet2", index=False)
    writer.save()

#Bảng 3
def bieu_do_tron2(ex):
    ga_1 = ex.pivot_table(index="Year", columns="Condition", values="Price", aggfunc=sum)
    print("Bảng doanh thu các năm theo tình trạng xe")
    print(ga_1)
    plt.pie([ga_1["Honda Certified"].sum(), ga_1["Used"].sum(), ga_1["New"].sum()],
            labels=["Honda Certified", "Used", "New"], autopct='%1.1f%%',
            startangle=15, shadow=True, explode=(0.1, 0, 0))
    plt.title("Biểu đồ phần trăm doanh thu theo tình trạng xe")
    plt.show()
    df = pd.read_excel("D:/Duy/honda_sample_2.xlsx", sheet_name="Sheet3")
    merged_df = pd.concat([df, ga_1], ignore_index=False)
    book = o.load_workbook("D:/Duy/honda_sample_2.xlsx")
    writer = pd.ExcelWriter("D:/Duy/honda_sample_2.xlsx", engine='openpyxl')
    writer.book = book
    merged_df.to_excel(writer, sheet_name="Sheet3", index=False)
    writer.save()

#Bảng 4
def bieu_do_cot3(ex):
    ga_2 = ex["Engine"].value_counts()
    print("Bảng tổng số xe bán được theo loại động cơ")
    print(ga_2)
    ga_2.sort_values().plot.barh(fontsize=5)
    plt.title("BIểu đồ tổng số xe bán được theo loại động cơ")
    plt.xlabel("Số lượng")
    plt.show()
    df = pd.read_excel("D:/Duy/honda_sample_2.xlsx", sheet_name="Sheet4")
    merged_df = pd.concat([df, ga_2], ignore_index=False)
    book = o.load_workbook("D:/Duy/honda_sample_2.xlsx")
    writer = pd.ExcelWriter("D:/Duy/honda_sample_2.xlsx", engine='openpyxl')
    writer.book = book
    merged_df.to_excel(writer, sheet_name="Sheet4", index=False)
    writer.save()

#Bảng 5
def bieu_do_cot2(ex):
    ga_1 = ex["Exterior_Color"].value_counts().sort_values()
    print("Bảng số xe bán được theo màu ngoại thất")
    print(ga_1)
    ga_1.plot.barh(fontsize=5)
    plt.title("Biểu đồ biễu diễn màu xe bán được nhiều nhất")
    plt.xlabel("Số lượng")
    plt.show()
    df = pd.read_excel("D:/Duy/honda_sample_2.xlsx", sheet_name="Sheet5")
    merged_df = pd.concat([df, ga_1], ignore_index=False)
    book = o.load_workbook("D:/Duy/honda_sample_2.xlsx")
    writer = pd.ExcelWriter("D:/Duy/honda_sample_2.xlsx", engine='openpyxl')
    writer.book = book
    merged_df.to_excel(writer, sheet_name="Sheet5", index=False)
    writer.save()






#Bảng 6
def histogram2(ex):
    print("Bảng số lượt xếp hạng")
    print(ex["Consumer_Rating"].value_counts())
    plt.hist(ex["Consumer_Rating"], bins=10, ec='b', fc='r', alpha=0.5, density=True, cumulative=True)
    plt.legend()
    plt.title("Biểu đồ tần số của mức xếp hạng")
    plt.xlabel("Mức xếp hạng")
    plt.ylabel("Số lượng")
    plt.show()



def cac_thong_ke(ex):
    a = ex[ex["Year"]==2022]
    pd.set_option('display.max_columns', None)  # Hiển thị tất cả các cột
    pd.set_option('display.max_rows', None)
    print("Tính các thống kê cơ bản theo cột 'Price' theo năm 2022:")
    print("Tổng doanh thu:", a["Price"].sum())
    print("Doanh thu lớn nhất trong các năm:", a["Price"].max())
    print("Doanh thu thấp nhất trong các năm:", a["Price"].min())
    print("Doanh thu trung bình của hãng:", a["Price"].mean())
    print("Phương sai doanh thu của hãng:", n.var(n.array(a["Price"])))
    print("Khoảng tứ phân vị:")
    print("Khoảng Q1:", a["Price"].quantile(0.25))
    print("Khoảng Q2 (trung vị):", a["Price"].quantile(0.5))
    print("Khoảng Q3:", a["Price"].quantile(0.75))
#D:/Baitap/BT_EXCEL/honda_sell_data.csv



if __name__ == '__main__':
    a = input("Nhập file cần xử lí:")
    ex = pd.read_csv(a, index_col=None)
    # ga = ex.pivot_table(index="Year", columns="Xếp loại", values="Consumer_Review_#", aggfunc=sum)
    # print(ga)
    # gA = ex.pivot_table(index="Year", values="Consumer_Review_#", aggfunc=sum)
    # print(gA)
    cac_thong_ke(ex)
    bieu_do_cot1(ex)
    bieu_do_cot2(ex)
    bieu_do_cot3(ex)
    bieu_do_duong(ex)
    bieu_do_tron2(ex)
    histogram2(ex)

